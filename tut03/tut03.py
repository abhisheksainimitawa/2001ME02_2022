#Help https://youtu.be/H37f_x4wAC0
def octant_longest_subsequence_count():
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time = [] #  new list for time
 v = []    #  new list for v
 u = []    #  new list for u
 w = []    #  new list for w
 vd = []    # list for difference of Vavg-v
 ud = []    # list for difference of Uavg-u
 wd = []    # list for difference of Wavg-w
 oct = []   # list for storing octant

 lwb=load_workbook("input_octant_longest_subsequence.xlsx")
 sheet=lwb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time.append(sheet.cell(row=r,column=1).value)
  u.append(sheet.cell(row=r,column=2).value)
  v.append(sheet.cell(row=r,column=3).value)
  w.append(sheet.cell(row=r,column=4).value)

 v_avg=np.mean(v, dtype=np.float64) # calculate  average of u
 u_avg=np.mean(u, dtype=np.float64) # calculate average of v
 w_avg=np.mean(w, dtype=np.float64) # calculate  average of w
 
 for r in range(2,n+1):
  ud.append(float(sheet.cell(row=r,column=2).value)-u_avg) # pushing  row[1]-Uavg in ud list
  vd.append(float(sheet.cell(row=r,column=3).value)-v_avg) # pushing row[2]-Vavg in vd list
  wd.append(float(sheet.cell(row=r,column=4).value)-w_avg)	# pushing  row[3]-Wavg in wd list 

 for i in range(0,n-1):# find octants
  if ((ud[i]>=0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(1)   # adding the octant in the oct list after checking the condition of octant 1 to calculate overall octant 1 value     
  elif((ud[i]>=0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-1)   # adding the octant in the oct list after checking the condition of octant -1 to calculate overall octant -1 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(2)   # adding the octant in the oct list after checking the condition of octant 2 to calculate overall octant 2 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-2)   # adding the octant in the oct list after checking the condition of octant -2 to calculate overall octant -2 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(4)   # adding the octant in the oct list after checking the condition of octant 4 to calculate overall octant 4 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]<0 )):
   oct.append(-4)    # adding the octant in the oct list after checking the condition of octant -4 to calculate overall octant -4 value   
  elif((ud[i]<0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(3)   # adding the octant in the oct list after checking the condition of octant 3 to calculate overall octant 3 value
  else:
   oct.append(-3)   # adding the octant in the oct list after checking the condition of octant -3 to calculate overall octant -3 value
 subs1=0 # defining a variable for storing longest susequece of octant 1
 subs2=0 #defining a variable for storing longest susequece of octant -1
 subs3=0 #defining a variable for storing longest susequece of octant 2
 subs4=0 #defining a variable for storing longest susequece of octant -2
 subs5=0 #defining a variable for storing longest susequece of octant 3
 subs6=0 #defining a variable for storing longest susequece of octant -3
 subs7=0 #defining a variable for storing longest susequece of octant 4
 subs8=0 #defining a variable for storing longest susequece of octant -4

 lsubs1=0 # defining a variable for storing longest subsequence length for octant 1
 lsubs2=0 # defining a variable for storing longest subsequence length for octant -1
 lsubs3=0 # defining a variable for storing longest subsequence length for octant 2
 lsubs4=0 # defining a variable for storing longest subsequence length for octant -2
 lsubs5=0 # defining a variable for storing longest subsequence length for octant 3
 lsubs6=0 # defining a variable for storing longest subsequence length for octant -3
 lsubs7=0 # defining a variable for storing longest subsequence length for octant 4
 lsubs8=0 # defining a variable for storing longest subsequence length for octant -4

 count1=0 # defining variable for count of octant 1
 count2=0 # defining variable for count of octant -1
 count3=0 # defining variable for count of octant 2
 count4=0 # defining variable for count of octant -2
 count5=0 # defining variable for count of octant 3
 count6=0 # defining variable for count of octant -3
 count7=0 # defining variable for count of octant 4
 count8=0 # defining variable for count of octant -4

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


octant_longest_subsequence_count()
